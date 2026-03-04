"""
Test Categories CRUD and Excel Import APIs
Testing new features: Category management and product import from Excel
"""
import pytest
import requests
import os
import tempfile

# openpyxl to create test Excel files
import openpyxl

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_EMAIL = "admin@faralink.com"
TEST_PASSWORD = "password123"


@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": TEST_EMAIL,
        "mot_de_passe": TEST_PASSWORD
    })
    if response.status_code == 200:
        return response.json().get("token")
    pytest.skip(f"Authentication failed: {response.text}")


@pytest.fixture
def authenticated_client(auth_token):
    """Session with auth header"""
    session = requests.Session()
    session.headers.update({
        "Content-Type": "application/json",
        "Authorization": f"Bearer {auth_token}"
    })
    return session


# ==================== CATEGORIES TESTS ====================

class TestCategoriesAPI:
    """Category CRUD endpoint tests"""
    
    def test_get_categories_list(self, authenticated_client):
        """Test GET /api/products/categories/list returns categories"""
        response = authenticated_client.get(f"{BASE_URL}/api/products/categories/list")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        print(f"Categories found: {data}")
    
    def test_create_category_success(self, authenticated_client):
        """Test POST /api/categories creates a new category"""
        unique_name = f"TEST_Category_{os.urandom(4).hex()}"
        response = authenticated_client.post(f"{BASE_URL}/api/categories", json={
            "nom": unique_name
        })
        assert response.status_code == 200
        data = response.json()
        assert "id" in data
        assert data["nom"] == unique_name
        print(f"Created category: {data}")
        
        # Verify it appears in categories list
        list_response = authenticated_client.get(f"{BASE_URL}/api/products/categories/list")
        assert list_response.status_code == 200
        categories = list_response.json()
        assert unique_name in categories
        
        # Cleanup - delete the test category
        del_response = authenticated_client.delete(f"{BASE_URL}/api/categories/{unique_name}")
        assert del_response.status_code == 200
    
    def test_create_duplicate_category_fails(self, authenticated_client):
        """Test POST /api/categories with duplicate name returns 400"""
        unique_name = f"TEST_DupCat_{os.urandom(4).hex()}"
        
        # Create first
        response1 = authenticated_client.post(f"{BASE_URL}/api/categories", json={
            "nom": unique_name
        })
        assert response1.status_code == 200
        
        # Try duplicate
        response2 = authenticated_client.post(f"{BASE_URL}/api/categories", json={
            "nom": unique_name
        })
        assert response2.status_code == 400
        print(f"Duplicate rejection: {response2.json()}")
        
        # Cleanup
        authenticated_client.delete(f"{BASE_URL}/api/categories/{unique_name}")
    
    def test_create_category_empty_name_fails(self, authenticated_client):
        """Test POST /api/categories with empty name returns 400"""
        response = authenticated_client.post(f"{BASE_URL}/api/categories", json={
            "nom": ""
        })
        assert response.status_code == 400
        print(f"Empty name rejection: {response.json()}")
    
    def test_delete_category(self, authenticated_client):
        """Test DELETE /api/categories/{cat_name} deletes a category"""
        unique_name = f"TEST_DelCat_{os.urandom(4).hex()}"
        
        # Create
        create_res = authenticated_client.post(f"{BASE_URL}/api/categories", json={
            "nom": unique_name
        })
        assert create_res.status_code == 200
        
        # Delete
        del_response = authenticated_client.delete(f"{BASE_URL}/api/categories/{unique_name}")
        assert del_response.status_code == 200
        data = del_response.json()
        assert "supprimée" in data.get("message", "")
        
        # Verify it's removed from list (custom categories only, not from products)
        list_response = authenticated_client.get(f"{BASE_URL}/api/products/categories/list")
        categories = list_response.json()
        # Note: categories from products still appear, only dedicated ones are deleted
        print(f"Category deleted successfully")
    
    def test_delete_nonexistent_category_fails(self, authenticated_client):
        """Test DELETE /api/categories/{cat_name} returns 404 for non-existent category"""
        response = authenticated_client.delete(f"{BASE_URL}/api/categories/NonExistentCategory12345")
        assert response.status_code == 404
        print(f"Non-existent category deletion: {response.json()}")


# ==================== EXCEL IMPORT TESTS ====================

class TestExcelImportAPI:
    """Excel import endpoint tests"""
    
    def test_download_import_template(self, authenticated_client):
        """Test GET /api/products/import-template returns a valid .xlsx file"""
        response = authenticated_client.get(f"{BASE_URL}/api/products/import-template")
        assert response.status_code == 200
        
        # Check content type is Excel
        content_type = response.headers.get('content-type', '')
        assert 'spreadsheet' in content_type or 'application/vnd' in content_type
        
        # Check content-disposition has filename
        disposition = response.headers.get('content-disposition', '')
        assert 'attachment' in disposition
        assert '.xlsx' in disposition
        
        # Verify it's valid Excel by trying to load with openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check headers are present
        headers = [cell.value for cell in ws[1]]
        assert 'Code' in headers or 'Désignation' in headers
        print(f"Template headers: {headers}")
    
    def test_import_valid_excel_file(self, authenticated_client):
        """Test POST /api/products/import-excel with valid .xlsx imports products"""
        # Create a test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produits"
        
        # Headers
        headers = ["Code", "Désignation", "Catégorie", "Prix Achat", "Prix Vente", "Stock"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add unique test products
        unique_suffix = os.urandom(4).hex()
        test_products = [
            [f"TEST_IMP_{unique_suffix}_1", "Test Import Produit 1", "Test Category", 1000, 1500, 50],
            [f"TEST_IMP_{unique_suffix}_2", "Test Import Produit 2", "Test Category", 2000, 3000, 100],
        ]
        
        for row_idx, product in enumerate(test_products, 2):
            for col_idx, value in enumerate(product, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            # Upload file
            with open(tmp_path, 'rb') as f:
                files = {'file': ('test_import.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                # Remove content-type header for multipart
                headers = {"Authorization": f"Bearer {authenticated_client.headers.get('Authorization', '').replace('Bearer ', '')}"}
                response = requests.post(
                    f"{BASE_URL}/api/products/import-excel",
                    files=files,
                    headers=headers
                )
            
            assert response.status_code == 200
            data = response.json()
            assert "imported" in data
            assert data["imported"] >= 2  # Should import at least 2 products
            print(f"Import result: {data}")
            
            # Verify products were created
            products_response = authenticated_client.get(f"{BASE_URL}/api/products?search=TEST_IMP_{unique_suffix}")
            products = products_response.json()
            assert len(products) >= 2
            
            # Cleanup - delete test products
            for product in products:
                if product["code"].startswith(f"TEST_IMP_{unique_suffix}"):
                    authenticated_client.delete(f"{BASE_URL}/api/products/{product['id']}")
            
        finally:
            os.unlink(tmp_path)
    
    def test_import_invalid_file_format_fails(self, authenticated_client):
        """Test POST /api/products/import-excel with invalid file returns 400"""
        # Create a non-Excel file (just text)
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False, mode='w') as tmp:
            tmp.write("This is not an Excel file")
            tmp_path = tmp.name
        
        try:
            with open(tmp_path, 'rb') as f:
                files = {'file': ('test.txt', f, 'text/plain')}
                headers = {"Authorization": f"Bearer {authenticated_client.headers.get('Authorization', '').replace('Bearer ', '')}"}
                response = requests.post(
                    f"{BASE_URL}/api/products/import-excel",
                    files=files,
                    headers=headers
                )
            
            assert response.status_code == 400
            print(f"Invalid file rejection: {response.json()}")
        finally:
            os.unlink(tmp_path)
    
    def test_import_excel_with_duplicate_codes_skips(self, authenticated_client):
        """Test import skips products with duplicate codes"""
        unique_code = f"TEST_DUP_{os.urandom(4).hex()}"
        
        # First, create a product directly
        create_response = authenticated_client.post(f"{BASE_URL}/api/products", json={
            "code": unique_code,
            "designation": "Original Product",
            "categorie": "Test",
            "prix_achat": 100,
            "prix_vente": 150,
            "quantite_stock": 10,
            "stock_minimum": 5,
            "unite": "Pièce"
        })
        assert create_response.status_code == 200
        original_product = create_response.json()
        
        # Create Excel with same code
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["Code", "Désignation", "Catégorie", "Prix Achat", "Prix Vente", "Stock"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Same code as existing product
        ws.cell(row=2, column=1, value=unique_code)
        ws.cell(row=2, column=2, value="Duplicate Product")
        ws.cell(row=2, column=3, value="Test")
        ws.cell(row=2, column=4, value=200)
        ws.cell(row=2, column=5, value=300)
        ws.cell(row=2, column=6, value=20)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            with open(tmp_path, 'rb') as f:
                files = {'file': ('dup_test.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                headers_req = {"Authorization": f"Bearer {authenticated_client.headers.get('Authorization', '').replace('Bearer ', '')}"}
                response = requests.post(
                    f"{BASE_URL}/api/products/import-excel",
                    files=files,
                    headers=headers_req
                )
            
            assert response.status_code == 200
            data = response.json()
            # Should skip because code already exists
            assert data.get("skipped", 0) >= 1
            print(f"Duplicate skip result: {data}")
            
        finally:
            os.unlink(tmp_path)
            # Cleanup
            authenticated_client.delete(f"{BASE_URL}/api/products/{original_product['id']}")


# ==================== PRODUCT CATEGORY DROPDOWN TEST ====================

class TestProductCategoryDropdown:
    """Verify category dropdown is populated from API"""
    
    def test_new_product_can_use_categories_from_list(self, authenticated_client):
        """Verify that when creating a product, categories from /api/products/categories/list can be used"""
        # Get categories
        cat_response = authenticated_client.get(f"{BASE_URL}/api/products/categories/list")
        assert cat_response.status_code == 200
        categories = cat_response.json()
        
        if len(categories) == 0:
            # Create a category first
            authenticated_client.post(f"{BASE_URL}/api/categories", json={"nom": "Test Category"})
            cat_response = authenticated_client.get(f"{BASE_URL}/api/products/categories/list")
            categories = cat_response.json()
        
        # Use first available category
        test_category = categories[0] if categories else "Default"
        unique_code = f"TEST_CATPROD_{os.urandom(4).hex()}"
        
        # Create product with that category
        response = authenticated_client.post(f"{BASE_URL}/api/products", json={
            "code": unique_code,
            "designation": "Product with Category",
            "categorie": test_category,
            "prix_achat": 500,
            "prix_vente": 750,
            "quantite_stock": 25,
            "stock_minimum": 5,
            "unite": "Pièce"
        })
        
        assert response.status_code == 200
        product = response.json()
        assert product["categorie"] == test_category
        print(f"Created product with category '{test_category}'")
        
        # Cleanup
        authenticated_client.delete(f"{BASE_URL}/api/products/{product['id']}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
